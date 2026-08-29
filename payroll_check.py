"""
payroll_check.py

Automates the accounts team's monthly salary-working validation:
  1. CTC cross-check          -- current month CTC vs previous month CTC (by Emp Id)
  2. Salary calculation check -- Basic/HRA/Conveyance/Special Allowance/Bonus/Gross/
                                  PF/ESI/PT recomputed from CTC + Salary Days, compared
                                  against what the pay register actually reports
  3. Bank details check       -- Account Number & IFSC in the pay register vs the
                                  Bank Master reference

This mirrors the formulas already used in the accounts team's Google Sheet
(Master Sheet / CTC / CTC - Previous Mon / Bank Master / PT / Check tabs), so the
numbers this produces should match what that sheet already shows -- just without
anyone having to VLOOKUP or eyeball anything by hand.

Explicitly NOT checked (per the accounts team's own process, since these aren't
comparable to CTC):
  - Variable Pay        -- actual/achievement-based, not fixed by CTC
  - Flexi components    -- depend on bills the employee submits that month
  - Salary Advance / Salary Advance deduction -- one-off, not CTC-driven
"""
import io
import re
from dataclasses import dataclass, field

import msoffcrypto
import openpyxl
import pandas as pd

# ----------------------------------------------------------------------
# Statutory constants (from the accounts team's own PT slab tables --
# update here if rates change)
# ----------------------------------------------------------------------
PF_RATE = 0.12
PF_CAP = 1800
ESI_RATE = 0.0075
ESI_WAGE_CEILING = 21000

PT_SLABS_CHENNAI = [  # (gross_from, gross_to_inclusive, pt_amount)
    (0, 3500, 0),
    (3501, 5000, 30),
    (5001, 7500, 71),
    (7501, 10000, 155),
    (10001, 12500, 171),
    (12501, float("inf"), 208),
]
PT_SLABS_OTHER = [  # Gujarat, used for any location that isn't Chennai
    (0, 5999, 0),
    (6000, 8999, 0),
    (9000, 11999, 0),
    (12000, float("inf"), 200),
]

DIFF_TOLERANCE = 1.0  # rupees -- ignore rounding noise below this


def lookup_pt(gross, location):
    slabs = PT_SLABS_CHENNAI if str(location).strip().lower() == "chennai" else PT_SLABS_OTHER
    for lo, hi, amt in slabs:
        if lo <= gross <= hi:
            return amt
    return slabs[-1][2]


# ----------------------------------------------------------------------
# File loading helpers
# ----------------------------------------------------------------------
def decrypt_if_needed(file_bytes: bytes, password: str | None) -> bytes:
    """Returns decrypted xlsx bytes if the file is password-protected,
    otherwise returns the original bytes unchanged."""
    buf = io.BytesIO(file_bytes)
    try:
        office_file = msoffcrypto.OfficeFile(buf)
        is_encrypted = office_file.is_encrypted()
    except Exception:
        is_encrypted = False

    if not is_encrypted:
        return file_bytes

    if not password:
        raise ValueError(
            "This file is password-protected but no password was provided."
        )
    buf.seek(0)
    office_file = msoffcrypto.OfficeFile(buf)
    office_file.load_key(password=password)
    out = io.BytesIO()
    office_file.decrypt(out)
    return out.getvalue()


def find_col(columns, *keywords):
    """Find the first column name containing all keywords (case-insensitive)."""
    for col in columns:
        val = str(col).strip().lower()
        if all(kw.lower() in val for kw in keywords):
            return col
    return None


def find_esi_column(columns):
    """Finds the ESI column while avoiding a false-positive substring match:
    the word "Designation" contains the letters "esi" (d-ESI-gnation), so a
    loose substring search for "esi" can wrongly grab the Designation column
    instead. Exact/near-exact match (spaces/dots stripped) is tried first."""
    for col in columns:
        normalized = re.sub(r"[\s.]+", "", str(col).strip().lower())
        if normalized == "esi":
            return col
    for col in columns:
        val = str(col).strip().lower()
        if "esi" in val and "designation" not in val:
            return col
    return None


def find_sheet(wb, *keywords):
    for name in wb.sheetnames:
        if all(kw.lower() in name.lower() for kw in keywords):
            return name
    return None


def read_sheet_as_df(xlsx_bytes, sheet_name):
    """Reads a sheet preserving exact cell types (so large account numbers
    stored as integers don't get corrupted by pandas' float64 coercion,
    which silently loses precision beyond ~15-16 significant digits)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    header = [str(h) if h is not None else f"Unnamed_{i}" for i, h in enumerate(rows[0])]
    data = rows[1:]
    df = pd.DataFrame(data, columns=header)
    return df


def clean_id(series):
    return series.astype(str).str.strip()


def clean_str(series):
    return series.astype(str).str.strip()


def normalize_account_number(series):
    def norm(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        if isinstance(v, float):
            v = str(int(round(v)))
        else:
            v = re.sub(r"\.0$", "", str(v).strip())
        # Strip leading zeros for comparison -- some sheets store account
        # numbers as text (preserving leading zeros) and others as numbers
        # (which can't have leading zeros), so "000901559228" and
        # "901559228" are the same account, not a mismatch.
        return v.lstrip("0") or "0"
    return series.apply(norm)


# ----------------------------------------------------------------------
# Loaders for each of the four inputs
# ----------------------------------------------------------------------
def load_payregister(xlsx_bytes):
    """Returns (salary_df, bank_df) from the raw pay register workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    master_sheet = find_sheet(wb, "master") or wb.sheetnames[0]
    bank_sheet = find_sheet(wb, "bank")
    wb.close()

    salary_df = read_sheet_as_df(xlsx_bytes, master_sheet)
    bank_df = read_sheet_as_df(xlsx_bytes, bank_sheet) if bank_sheet else None
    return salary_df, bank_df


def load_ctc_master(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    sheet = find_sheet(wb, "ctc") or wb.sheetnames[0]
    wb.close()
    return read_sheet_as_df(xlsx_bytes, sheet)


def load_bank_master(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    sheet = find_sheet(wb, "bank") or wb.sheetnames[0]
    wb.close()
    return read_sheet_as_df(xlsx_bytes, sheet)


# ----------------------------------------------------------------------
# Main check
# ----------------------------------------------------------------------
@dataclass
class CheckResults:
    ctc_changes: pd.DataFrame
    salary_exceptions: pd.DataFrame
    bank_mismatches: pd.DataFrame
    not_in_ctc: pd.DataFrame
    not_in_bank_master: pd.DataFrame
    total_employees: int
    days_in_month: float = 30
    warnings: list = field(default_factory=list)


def run_checks(salary_df, bank_df, ctc_df, prev_ctc_df, bank_master_df) -> CheckResults:
    warnings = []

    # --- column resolution (tolerant to minor header wording changes) ---
    c_emp = find_col(salary_df.columns, "emp", "id")
    c_name = find_col(salary_df.columns, "associate", "name") or find_col(salary_df.columns, "name")
    c_days = find_col(salary_df.columns, "salary", "days")
    c_location = find_col(salary_df.columns, "location")
    c_r_basic = find_col(salary_df.columns, "basic", "pay") or find_col(salary_df.columns, "basic")
    c_r_hra = find_col(salary_df.columns, "earned", "hra") or find_col(salary_df.columns, "hra")
    c_r_conv = find_col(salary_df.columns, "earned", "conveyance") or find_col(salary_df.columns, "conveyance")
    c_r_bonus = find_col(salary_df.columns, "earned", "bonus") or find_col(salary_df.columns, "bonus")
    c_r_spl = find_col(salary_df.columns, "earned", "spl", "allowance") or find_col(salary_df.columns, "spl", "allowance")
    c_r_gross = find_col(salary_df.columns, "gross", "earning")
    c_r_pf = find_col(salary_df.columns, "provident", "fund")
    c_r_esi = find_esi_column(salary_df.columns)
    c_r_pt = find_col(salary_df.columns, "professional", "tax")

    c_lop_basic = find_col(salary_df.columns, "lop", "credit", "basic")
    c_lop_hra = find_col(salary_df.columns, "lop", "credit", "hra")
    c_lop_conv = find_col(salary_df.columns, "lop", "credit", "conveyance")
    c_lop_bonus = find_col(salary_df.columns, "lop", "credit", "bonus")
    c_lop_spl = find_col(salary_df.columns, "lop", "credit", "spl")
    c_incentive = find_col(salary_df.columns, "incentive")
    c_night_shift = find_col(salary_df.columns, "night", "shift")
    c_arrear = find_col(salary_df.columns, "arrear") if not find_col(salary_df.columns, "arrear", "provident") else find_col(salary_df.columns, "arrear")
    c_other_earn = find_col(salary_df.columns, "other", "earning")
    c_med_reimb = find_col(salary_df.columns, "medical", "reimb")
    c_tel_reimb = find_col(salary_df.columns, "telephone", "reimb")
    c_fuel_reimb = find_col(salary_df.columns, "fuel", "reimb")
    c_lta_reimb = find_col(salary_df.columns, "lta", "reimb")
    c_med_tax = find_col(salary_df.columns, "medical", "taxable")
    c_tel_tax = find_col(salary_df.columns, "telephone", "taxable")
    c_fuel_tax = find_col(salary_df.columns, "fuel", "taxable")
    c_lta_tax = find_col(salary_df.columns, "lta", "taxable")
    c_variable = find_col(salary_df.columns, "variable", "pay")
    c_sunday_ot = find_col(salary_df.columns, "sunday", "ot")
    c_ot = find_col(salary_df.columns, "ot") if not c_sunday_ot else find_col(
        [c for c in salary_df.columns if c != c_sunday_ot], "ot"
    )

    c_ctc_emp = find_col(ctc_df.columns, "emp", "id")
    c_ctc_fixed = find_col(ctc_df.columns, "fixed", "ctc")
    c_ctc_basic = find_col(ctc_df.columns, "basic")
    c_ctc_hra = find_col(ctc_df.columns, "hra")
    c_ctc_conv = find_col(ctc_df.columns, "conveyance")
    c_ctc_bonus = find_col(ctc_df.columns, "bonus")
    c_ctc_spl = find_col(ctc_df.columns, "special", "allowance", "minus") or find_col(
        ctc_df.columns, "special", "allowance"
    )
    c_ctc_location = find_col(ctc_df.columns, "location")
    c_ctc_name = find_col(ctc_df.columns, "associate", "name") or find_col(ctc_df.columns, "name")
    c_ctc_designation = find_col(ctc_df.columns, "designation")
    c_ctc_doj = find_col(ctc_df.columns, "doj") or find_col(ctc_df.columns, "date", "of", "join")

    c_prev_emp = find_col(prev_ctc_df.columns, "emp", "id")
    c_prev_fixed = find_col(prev_ctc_df.columns, "fixed", "ctc")
    c_prev_basic = find_col(prev_ctc_df.columns, "basic")
    c_prev_hra = find_col(prev_ctc_df.columns, "hra")
    c_prev_conv = find_col(prev_ctc_df.columns, "conveyance")
    c_prev_bonus = find_col(prev_ctc_df.columns, "bonus")
    c_prev_spl = find_col(prev_ctc_df.columns, "special", "allowance", "minus") or find_col(
        prev_ctc_df.columns, "special", "allowance"
    )

    required = {
        "Emp Id (pay register)": c_emp, "Salary Days": c_days,
        "Fixed CTC (current)": c_ctc_fixed, "Emp Id (CTC)": c_ctc_emp,
        "Fixed CTC (previous)": c_prev_fixed, "Emp Id (previous CTC)": c_prev_emp,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError(f"Could not find required column(s): {', '.join(missing)}")

    df = salary_df.copy()
    df = df[df[c_emp].notna()].copy()
    df["_emp_id"] = clean_id(df[c_emp])
    name_col = df[c_name] if c_name else df[c_emp]

    def num(col):
        if col is None or col not in df.columns:
            return pd.Series(0.0, index=df.index)
        return pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    salary_days = num(c_days)
    location = df[c_location] if c_location else pd.Series("", index=df.index)

    # ---------------- CTC cross-check ----------------
    ctc = ctc_df.copy()
    ctc["_emp_id"] = clean_id(ctc[c_ctc_emp])
    ctc = ctc.drop_duplicates("_emp_id", keep="first").set_index("_emp_id")

    prev = prev_ctc_df.copy()
    prev["_emp_id"] = clean_id(prev[c_prev_emp])
    prev = prev.drop_duplicates("_emp_id", keep="first").set_index("_emp_id")

    all_emp_ids = df["_emp_id"].drop_duplicates()
    cur_fixed_map = all_emp_ids.map(ctc[c_ctc_fixed]) if c_ctc_fixed in ctc.columns else pd.Series(dtype=float)
    prev_fixed_map = all_emp_ids.map(prev[c_prev_fixed]) if c_prev_fixed in prev.columns else pd.Series(dtype=float)

    not_in_ctc = all_emp_ids[cur_fixed_map.isna().values].tolist()

    ctc_change_rows = []
    for emp_id, cur_fixed, prev_fixed in zip(all_emp_ids, cur_fixed_map, prev_fixed_map):
        if pd.isna(cur_fixed):
            continue  # not in current CTC at all -- handled separately (not_in_ctc)
        if pd.isna(prev_fixed):
            # In current month's CTC but not last month's -- a new joiner.
            # Report this explicitly with their details rather than silently
            # skipping it, since a new employee showing up is worth a human
            # glance even though it's not a "change" requiring approval.
            ctc_row = ctc.loc[emp_id] if emp_id in ctc.index else None
            ctc_change_rows.append({
                "Emp Id": emp_id,
                "Name": ctc_row[c_ctc_name] if (ctc_row is not None and c_ctc_name) else "",
                "Status": "New Employee",
                "Designation": ctc_row[c_ctc_designation] if (ctc_row is not None and c_ctc_designation) else "",
                "Location": ctc_row[c_ctc_location] if (ctc_row is not None and c_ctc_location) else "",
                "Date of Joining": ctc_row[c_ctc_doj] if (ctc_row is not None and c_ctc_doj) else "",
                "Previous Fixed CTC": None,
                "Current Fixed CTC": cur_fixed,
                "Difference": None,
                "Needs HR Approval": "No -- new joiner",
            })
            continue
        if abs(float(cur_fixed) - float(prev_fixed)) > DIFF_TOLERANCE:
            ctc_change_rows.append({
                "Emp Id": emp_id,
                "Name": ctc.loc[emp_id, c_ctc_name] if (emp_id in ctc.index and c_ctc_name) else "",
                "Status": "CTC Changed",
                "Designation": ctc.loc[emp_id, c_ctc_designation] if (emp_id in ctc.index and c_ctc_designation) else "",
                "Location": ctc.loc[emp_id, c_ctc_location] if (emp_id in ctc.index and c_ctc_location) else "",
                "Date of Joining": ctc.loc[emp_id, c_ctc_doj] if (emp_id in ctc.index and c_ctc_doj) else "",
                "Previous Fixed CTC": prev_fixed,
                "Current Fixed CTC": cur_fixed,
                "Difference": float(cur_fixed) - float(prev_fixed),
                "Needs HR Approval": "Yes",
            })
    ctc_changes = pd.DataFrame(ctc_change_rows)
    if not ctc_changes.empty:
        column_order = [
            "Emp Id", "Name", "Status", "Designation", "Location", "Date of Joining",
            "Previous Fixed CTC", "Current Fixed CTC", "Difference", "Needs HR Approval",
        ]
        ctc_changes = ctc_changes[column_order]

    c_lop_days = find_col(salary_df.columns, "loss", "of", "pay", "days")

    # ---------------- Salary calculation check ----------------
    def ctc_val(col):
        if col not in ctc.columns:
            return pd.Series(0.0, index=df.index)
        s = df["_emp_id"].map(ctc[col])
        return pd.to_numeric(s, errors="coerce").fillna(0.0)

    ctc_basic = ctc_val(c_ctc_basic)
    ctc_hra = ctc_val(c_ctc_hra)
    ctc_conv = ctc_val(c_ctc_conv)
    ctc_bonus = ctc_val(c_ctc_bonus)
    ctc_spl = ctc_val(c_ctc_spl)

    # Days in the salary month: derived from employees with zero LOP, whose
    # Salary Days should equal the full length of the month (28/29/30/31).
    # This is NOT hardcoded to 30 -- payroll here prorates against the actual
    # number of days in the calendar month.
    days_in_month = 30
    if c_lop_days:
        lop_vals = num(c_lop_days)
        full_month_days = pd.to_numeric(df.loc[lop_vals.values == 0, c_days], errors="coerce").dropna()
        if not full_month_days.empty:
            mode = full_month_days.mode()
            if not mode.empty:
                days_in_month = float(mode.iloc[0])

    if salary_days.max() <= 0:
        raise ValueError(
            f"The '{c_days}' column in the Pay Register is 0 (or blank) for every "
            "employee, so proration can't be computed and every 'Expected' figure "
            "would come out as 0. This looks like a data issue in the Pay Register "
            "itself -- check with whoever generated it that the Salary Days and "
            "Loss of Pay Days columns are filled in correctly for this month, "
            "then re-upload."
        )

    factor = salary_days / days_in_month
    exp_basic = ctc_basic * factor
    exp_hra = ctc_hra * factor
    exp_conv = ctc_conv * factor
    exp_bonus = ctc_bonus * factor
    exp_spl = ctc_spl * factor

    lop_basic = num(c_lop_basic)
    lop_hra = num(c_lop_hra)
    lop_conv = num(c_lop_conv)
    lop_bonus = num(c_lop_bonus)
    lop_spl = num(c_lop_spl)
    incentive = num(c_incentive)
    night_shift = num(c_night_shift)
    arrear = num(c_arrear)
    other_earn = num(c_other_earn)
    med_reimb = num(c_med_reimb)
    tel_reimb = num(c_tel_reimb)
    fuel_reimb = num(c_fuel_reimb)
    lta_reimb = num(c_lta_reimb)
    med_tax = num(c_med_tax)
    tel_tax = num(c_tel_tax)
    fuel_tax = num(c_fuel_tax)
    lta_tax = num(c_lta_tax)
    variable = num(c_variable)
    sunday_ot = num(c_sunday_ot)
    ot = num(c_ot)
    others = arrear + other_earn + night_shift

    exp_gross = (
        exp_basic + exp_hra + exp_conv + exp_spl + exp_bonus
        + lop_basic + lop_hra + lop_conv + lop_bonus + lop_spl
        + incentive
        + med_reimb + tel_reimb + fuel_reimb + lta_reimb
        + med_tax + tel_tax + fuel_tax + lta_tax
        + variable + others + ot + sunday_ot
    )

    # --- New CTC breakup model: PF, ESI (and Bonus) are based on wages,
    # defined as Basic + Special Allowance -- recomputed FRESH from this
    # month's actual earned amounts (not simply prorating an already-capped
    # full-month figure, since the PF cap is nonlinear across partial months).
    wages_new_model = exp_basic + exp_spl + lop_basic + lop_spl

    # Some employees have individually opted for uncapped PF (12% of Basic
    # alone, no ceiling) instead of the standard capped formula. We can't
    # derive who from the wage alone -- but CTC Master already encodes this
    # decision: if their full-month PF Employee figure exceeds the cap, they
    # must be on the uncapped basis.
    c_ctc_pf_emp = find_col(ctc_df.columns, "pf", "employee")
    ctc_pf_full_month = ctc_val(c_ctc_pf_emp) if c_ctc_pf_emp else pd.Series(0.0, index=df.index)
    is_uncapped_pf = ctc_pf_full_month > (PF_CAP + DIFF_TOLERANCE)

    pf_capped = (wages_new_model * PF_RATE).clip(upper=PF_CAP)
    pf_uncapped = exp_basic * PF_RATE
    pf_expected = pf_uncapped.where(is_uncapped_pf, pf_capped)

    # ESI eligibility itself is decided at the CTC level (based on the
    # employee's full monthly wage) and does not flip on and off due to a
    # single month's Loss-of-Pay -- so an employee CTC Master has already
    # excluded from ESI stays excluded even if this month's prorated wage
    # happens to dip below the ceiling. Only the AMOUNT for already-eligible
    # employees is recomputed fresh from this month's actual wage.
    c_ctc_esi_emp = find_col(ctc_df.columns, "esi", "employee")
    ctc_esi_full_month = ctc_val(c_ctc_esi_emp) if c_ctc_esi_emp else pd.Series(0.0, index=df.index)
    esi_eligible = ctc_esi_full_month > DIFF_TOLERANCE

    esi_fresh = wages_new_model.apply(lambda w: w * ESI_RATE if w < ESI_WAGE_CEILING else 0.0)
    esi_expected = esi_fresh.where(esi_eligible, 0.0)

    reported_gross = num(c_r_gross)
    pt_expected = pd.Series(
        [lookup_pt(g_, loc) for g_, loc in zip(reported_gross, location)],
        index=df.index,
    )

    result = pd.DataFrame({
        "Emp Id": df["_emp_id"],
        "Name": name_col,
        "Salary Days": salary_days,
        "Reported Basic": num(c_r_basic), "Expected Basic": exp_basic,
        "Reported HRA": num(c_r_hra), "Expected HRA": exp_hra,
        "Reported Conveyance": num(c_r_conv), "Expected Conveyance": exp_conv,
        "Reported Special Allowance": num(c_r_spl), "Expected Special Allowance": exp_spl,
        "Reported Bonus": num(c_r_bonus), "Expected Bonus": exp_bonus,
        "Reported Gross": reported_gross, "Expected Gross": exp_gross,
        "Reported PF": num(c_r_pf), "Expected PF": pf_expected,
        "Reported ESI": num(c_r_esi), "Expected ESI": esi_expected,
        "Reported PT": num(c_r_pt), "Expected PT": pt_expected,
    })

    for label in ["Basic", "HRA", "Conveyance", "Special Allowance", "Bonus", "Gross", "PF", "ESI", "PT"]:
        result[f"{label} Diff"] = result[f"Reported {label}"] - result[f"Expected {label}"]

    diff_cols = [f"{l} Diff" for l in ["Basic", "HRA", "Conveyance", "Special Allowance", "Bonus", "Gross", "PF", "ESI", "PT"]]
    # NaN comparisons are always False (e.g. NaN > 1 is False, not an error),
    # so a row where every diff is NaN -- which happens if this employee's
    # own Salary Days is 0/blank, even if everyone else's data is fine --
    # would otherwise silently pass this check instead of being flagged as
    # the computation failure it actually is.
    has_exception = result[diff_cols].abs().gt(DIFF_TOLERANCE).any(axis=1) | result[diff_cols].isna().any(axis=1)
    salary_exceptions = result[has_exception].copy()

    # ---------------- Bank details check ----------------
    bank_mismatches = pd.DataFrame()
    not_in_bank_master = pd.DataFrame()
    if bank_df is not None and bank_master_df is not None:
        c_b_emp = find_col(bank_df.columns, "emp", "id")
        c_b_name = find_col(bank_df.columns, "associate", "name") or find_col(bank_df.columns, "name")
        c_b_acc = find_col(bank_df.columns, "account", "number")
        c_b_ifsc = find_col(bank_df.columns, "ifsc")

        c_bm_emp = find_col(bank_master_df.columns, "emp", "id")
        c_bm_acc = find_col(bank_master_df.columns, "account", "number")
        c_bm_ifsc = find_col(bank_master_df.columns, "ifsc")

        if all([c_b_emp, c_b_acc, c_b_ifsc, c_bm_emp, c_bm_acc, c_bm_ifsc]):
            bm = bank_master_df.copy()
            bm = bm[bm[c_bm_emp].notna()].copy()
            bm["_emp_id"] = clean_id(bm[c_bm_emp])
            bm = bm.drop_duplicates("_emp_id", keep="first").set_index("_emp_id")

            bd = bank_df.copy()
            bd = bd[bd[c_b_emp].notna()].copy()
            bd["_emp_id"] = clean_id(bd[c_b_emp])

            master_acc = bd["_emp_id"].map(bm[c_bm_acc])
            master_ifsc = bd["_emp_id"].map(bm[c_bm_ifsc])

            reported_acc = normalize_account_number(bd[c_b_acc])
            reported_ifsc = clean_str(bd[c_b_ifsc].astype(str))
            master_acc_clean = normalize_account_number(master_acc)
            master_ifsc_clean = clean_str(master_ifsc.astype(str))

            missing_mask = master_acc.isna()
            acc_mismatch = (~missing_mask) & (reported_acc != master_acc_clean)
            ifsc_mismatch = (~missing_mask) & (reported_ifsc != master_ifsc_clean)

            bank_check = pd.DataFrame({
                "Emp Id": bd["_emp_id"],
                "Name": bd[c_b_name] if c_b_name else bd["_emp_id"],
                "Reported Account No": bd[c_b_acc],
                "Bank Master Account No": master_acc,
                "Reported IFSC": bd[c_b_ifsc],
                "Bank Master IFSC": master_ifsc,
            })
            not_in_bank_master = bank_check[missing_mask].copy()
            bank_mismatches = bank_check[(acc_mismatch | ifsc_mismatch) & ~missing_mask].copy()
        else:
            warnings.append(
                "Could not find Account Number / IFSC / Emp Id columns in one of the "
                "bank sheets -- bank check skipped."
            )
    else:
        warnings.append("Bank sheet or Bank Master not provided -- bank check skipped.")

    return CheckResults(
        ctc_changes=ctc_changes,
        salary_exceptions=salary_exceptions,
        bank_mismatches=bank_mismatches,
        not_in_ctc=pd.DataFrame({"Emp Id": not_in_ctc}),
        not_in_bank_master=not_in_bank_master,
        total_employees=len(df),
        days_in_month=days_in_month,
        warnings=warnings,
    )


# ----------------------------------------------------------------------
# Excel report writer
# ----------------------------------------------------------------------
def write_report(results: CheckResults) -> bytes:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bad_fill = PatternFill(start_color="FDE7EC", end_color="FDE7EC", fill_type="solid")
    ok_fill = PatternFill(start_color="E7F6EC", end_color="E7F6EC", fill_type="solid")

    ws_summary.append(["Payroll Validation Summary"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Total employees in pay register", results.total_employees])
    ws_summary.append(["Days in month used for proration (auto-detected)", results.days_in_month])
    ws_summary.append(["CTC changes requiring HR approval", len(results.ctc_changes)])
    ws_summary.append(["Salary calculation exceptions", len(results.salary_exceptions)])
    ws_summary.append(["Bank detail mismatches", len(results.bank_mismatches)])
    ws_summary.append(["Employees not found in CTC master", len(results.not_in_ctc)])
    ws_summary.append(["Employees not found in Bank Master", len(results.not_in_bank_master)])
    if results.warnings:
        ws_summary.append([])
        ws_summary.append(["Warnings:"])
        for w in results.warnings:
            ws_summary.append([w])
    ws_summary.column_dimensions["A"].width = 42
    ws_summary.column_dimensions["B"].width = 14

    def write_df_sheet(name, dframe, diff_cols=None):
        ws = wb.create_sheet(name)
        if dframe is None or dframe.empty:
            ws.append([f"No exceptions -- all clear."])
            ws["A1"].fill = ok_fill
            ws["A1"].font = Font(bold=True, color="1B7A3D")
            return
        ws.append(list(dframe.columns))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for _, row in dframe.iterrows():
            ws.append(list(row))
        if diff_cols:
            for col_name in diff_cols:
                if col_name not in dframe.columns:
                    continue
                col_idx = list(dframe.columns).index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for r in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{r}"]
                    try:
                        if abs(float(cell.value)) > DIFF_TOLERANCE:
                            cell.fill = bad_fill
                    except (TypeError, ValueError):
                        pass
        for col_idx, col_name in enumerate(dframe.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(30, len(str(col_name)) + 4))
        ws.freeze_panes = "A2"

    write_df_sheet("CTC Changes", results.ctc_changes, diff_cols=["Difference"])
    write_df_sheet(
        "Salary Exceptions",
        results.salary_exceptions,
        diff_cols=[c for c in results.salary_exceptions.columns if c.endswith("Diff")] if not results.salary_exceptions.empty else None,
    )
    write_df_sheet("Bank Mismatches", results.bank_mismatches)
    write_df_sheet("Not in CTC Master", results.not_in_ctc)
    write_df_sheet("Not in Bank Master", results.not_in_bank_master)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
